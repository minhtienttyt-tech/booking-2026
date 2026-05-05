import openpyxl
import json
import os

file_path = 'Booking Xuất VAT 2026 (1).xlsx'
output_path = '.agent/scratch/excel_structure.json'

if not os.path.exists('.agent/scratch'):
    os.makedirs('.agent/scratch')

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    result = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Read first 20 rows of each sheet
        rows = []
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
            clean_row = []
            for cell in row:
                if cell is not None and hasattr(cell, 'isoformat'):
                    clean_row.append(cell.isoformat())
                else:
                    clean_row.append(cell)
            rows.append(clean_row)
        result[sheet_name] = rows
        
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"Successfully wrote structure to {output_path}")
except Exception as e:
    print(f"Error: {e}")
